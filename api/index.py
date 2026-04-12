from flask import Flask, request, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import Workbook, load_workbook
import os
import uuid 

app = Flask(__name__)
# IMPORTANT: This allows your frontend to talk to this backend
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))


FILE_NAME = os.path.join(BASE_DIR, 'expenses.xlsx')
META_FILE_NAME = os.path.join(BASE_DIR, 'meta_data.xlsx')
COLUMNS = ['ID', 'Date', 'Type', 'Category', 'Description', 'Amount']

def init_excel():
    if not os.path.exists(META_FILE_NAME):
        meta_df = pd.DataFrame([{'Starting Balance': 0.0}])
        with pd.ExcelWriter(META_FILE_NAME, engine='openpyxl') as writer:
            meta_df.to_excel(writer, index=False, sheet_name='Config')

    if not os.path.exists(FILE_NAME):
        df = pd.DataFrame(columns=COLUMNS)
        with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Expenses')
        
        wb = load_workbook(FILE_NAME)
        ws = wb['Expenses']
        for cell in ws[1]:
            cell.font = Font(bold=True)
            
        for row in ws.iter_rows(min_col=6, max_col=6, min_row=2):
            for cell in row:
                cell.number_format = '"₹"#,##0.00'
                
        wb.save(FILE_NAME)
    else:
        df = pd.read_excel(FILE_NAME, engine='openpyxl')
        if 'Type' not in df.columns:
            df.insert(2, 'Type', 'Expense')
            with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Expenses')
            
            wb = load_workbook(FILE_NAME)
            ws = wb['Expenses']
            for cell in ws[1]:
                cell.font = Font(bold=True)
                
            for row in ws.iter_rows(min_col=6, max_col=6, min_row=2):
                for cell in row:
                    cell.number_format = '"₹"#,##0.00'
                    
            wb.save(FILE_NAME)

@app.route('/api/expenses', methods=['GET'])
def get_expenses():
    if not os.path.exists(FILE_NAME):
        return jsonify([])
    
    try:
        df = pd.read_excel(FILE_NAME, engine='openpyxl')
        # We need to ensure we don't send NaN as JSON
        df = df.where(pd.notnull(df), None)
        expenses = df.to_dict('records')
        
        formatted_expenses = [
            {
                "id": str(e.get('ID')) if e.get('ID') is not None else str(uuid.uuid4()),
                "date": str(e.get('Date'))[:10] if e.get('Date') else '',
                "type": e.get('Type', 'Expense'),
                "category": e.get('Category'),
                "description": e.get('Description'),
                "amount": e.get('Amount')
            }
            for e in expenses
        ]
        # Reverse to show latest expenses at the top, matching UI expectations
        return jsonify(list(reversed(formatted_expenses)))
    except Exception as e:
        print("Error reading excel:", e)
        return jsonify([])

@app.route('/api/expenses', methods=['POST'])
def add_expense():
    init_excel()
    data = request.json
    
    new_id = str(uuid.uuid4())
    
    wb = load_workbook(FILE_NAME)
    ws = wb['Expenses']
    
    amount_val = float(data.get('amount', 0))
    ws.append([
        new_id, 
        data.get('date'), 
        data.get('type', 'Expense'),
        data.get('category'), 
        data.get('description'), 
        amount_val
    ])
    
    # Set currency format for the newly added amount (column 6)
    new_row = ws.max_row
    ws.cell(row=new_row, column=6).number_format = '"₹"#,##0.00'
    
    wb.save(FILE_NAME)
    
    return jsonify({
        "id": new_id,
        "date": data.get('date'),
        "type": data.get('type', 'Expense'),
        "category": data.get('category'),
        "description": data.get('description'),
        "amount": amount_val
    })

@app.route('/api/expenses/<expense_id>', methods=['DELETE'])
def delete_expense(expense_id):
    if not os.path.exists(FILE_NAME):
        return jsonify({"success": False, "error": "File not found"}), 404
        
    wb = load_workbook(FILE_NAME)
    ws = wb['Expenses']
    
    row_to_delete = None
    for idx, row in enumerate(ws.iter_rows(min_col=1, max_col=1, min_row=2), start=2):
        if str(row[0].value) == expense_id:
            row_to_delete = idx
            break
            
    if row_to_delete:
        ws.delete_rows(row_to_delete)
        wb.save(FILE_NAME)
        return jsonify({"success": True})
        
    return jsonify({"success": False, "error": "Not found"}), 404

@app.route('/api/meta', methods=['GET'])
def get_meta():
    if not os.path.exists(META_FILE_NAME):
        return jsonify({"startingBalance": 0})
    try:
        df = pd.read_excel(META_FILE_NAME, engine='openpyxl')
        balance = float(df.iloc[0]['Starting Balance']) if not df.empty else 0.0
        return jsonify({"startingBalance": balance})
    except Exception as e:
        print("Error reading meta data:", e)
        return jsonify({"startingBalance": 0})

@app.route('/api/meta', methods=['POST'])
def update_meta():
    data = request.json
    balance = float(data.get('startingBalance', 0))
    meta_df = pd.DataFrame([{'Starting Balance': balance}])
    with pd.ExcelWriter(META_FILE_NAME, engine='openpyxl') as writer:
        meta_df.to_excel(writer, index=False, sheet_name='Config')
    return jsonify({"startingBalance": balance})

if __name__ == '__main__':
    init_excel()
    app.run(port=5000, debug=True)
